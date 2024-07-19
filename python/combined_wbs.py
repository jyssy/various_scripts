import os
import pandas as pd
from openpyxl import load_workbook

# Get the current working directory
input_directory = os.getcwd()
output_file = 'combined_workbook.xlsx'

# Create a new Excel writer object
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Loop through each file in the directory
    for filename in os.listdir(input_directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_directory, filename)
            try:
                # Load the workbook
                workbook = load_workbook(file_path)
                # Loop through each sheet in the workbook
                for sheet_name in workbook.sheetnames:
                    # Read the sheet into a DataFrame
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    # Write the DataFrame to the new workbook
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                print(f"Error processing {file_path}: {e}")

print(f"All sheets have been combined into {output_file}")