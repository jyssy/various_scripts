import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

wb = Workbook()
# ws = wb.active
ws1 = wb.create_sheet('yes/no', 0)

yayornay = pd.read_csv('yayornay.csv', 'r', header=None)  #  the actual yay or nay counts
yayornay = dataframe_to_rows(yayornay)
for r_idx, row in enumerate(yayornay, 1):
    for c_idx, value in enumerate(row, 1):
         ws1.cell(row=r_idx, column=c_idx, value=value)


what = input('What year is it? ')

wb.save(what + '.FASTSurveyData.xlsx')